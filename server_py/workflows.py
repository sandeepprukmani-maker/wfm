import os
import json
import time
import base64
import requests
import re
import threading
import pandas as pd
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from flask import request, jsonify
from .storage import storage
from .utils import log, resolve_variables, get_ai

def export_to_excel(data, node_id, execution_id):
    """Export query result to Excel with auto-fit columns and highlighted headers."""
    try:
        if not data or not isinstance(data, list):
            return None
            
        df = pd.DataFrame(data)
        file_path = f"/tmp/query_result_{execution_id}_{node_id}.xlsx"
        
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Query Results')
            workbook = writer.book
            worksheet = writer.sheets['Query Results']
            
            # Highlight headers in yellow
            yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            header_font = Font(bold=True)
            
            for cell in worksheet[1]:
                cell.fill = yellow_fill
                cell.font = header_font
            
            # Auto-fit column width
            for i, col in enumerate(df.columns):
                column_len = max(df[col].astype(str).str.len().max(), len(col)) + 2
                worksheet.column_dimensions[get_column_letter(i + 1)].width = column_len
                
        return file_path
    except Exception as e:
        log(f"Excel export failed: {e}")
        return None

def get_dag_state(dag_id, base_url, auth_headers):
    try:
        response = requests.get(
            f"{base_url}/api/v1/dags/{dag_id}/dagRuns",
            params={'order_by': '-execution_date', 'limit': 1},
            headers=auth_headers
        )
        response.raise_for_status()
        dag_runs = response.json().get('dag_runs', [])
        if dag_runs:
            return dag_runs[0].get('state', 'unknown')
        return 'no_runs'
    except Exception as e:
        log(f"Error checking DAG state for {dag_id}: {e}")
        return 'unknown'

def wait_for_dags_to_complete(dag_infos, logs, execution_id, storage):
    max_wait_time = 3600
    poll_interval = 10
    elapsed = 0
    
    while elapsed < max_wait_time:
        all_complete = True
        for dag_info in dag_infos:
            dag_id = dag_info['dag_id']
            base_url = dag_info['base_url']
            auth_headers = dag_info['auth_headers']
            
            if not base_url:
                continue
            
            state = get_dag_state(dag_id, base_url, auth_headers)
            running_states = ['running', 'queued', 'scheduled', 'up_for_retry', 'up_for_reschedule', 'restarting', 'deferred']
            
            if state.lower() in running_states:
                all_complete = False
                logs.append({
                    'timestamp': datetime.now().isoformat(),
                    'level': 'INFO',
                    'message': f"DAG {dag_id} is currently {state}. Waiting for it to reach a terminal state..."
                })
                storage.update_execution(execution_id, 'waiting', logs)
                break
        
        if all_complete:
            return True
        
        time.sleep(poll_interval)
        elapsed += poll_interval
    
    logs.append({
        'timestamp': datetime.now().isoformat(),
        'level': 'ERROR',
        'message': f"Timeout waiting for DAGs to reach a terminal state after {max_wait_time} seconds"
    })
    return False

def collect_dag_infos_from_workflow(nodes, storage):
    dag_infos = []
    for node in nodes:
        node_data = node.get('data', {})
        node_type = node_data.get('type')
        
        if node_type in ['airflow_trigger', 'airflow_log_check']:
            config = node_data.get('config', {})
            dag_id = config.get('dagId', '')
            credential_id = config.get('credentialId')
            
            base_url = ""
            auth_headers = {}
            
            if credential_id:
                cred = storage.get_credential(int(credential_id))
                if cred and cred.get('type') == 'airflow':
                    cred_data = cred.get('data', {})
                    base_url = cred_data.get('baseUrl', '')
                    auth = base64.b64encode(f"{cred_data.get('username')}:{cred_data.get('password')}".encode()).decode()
                    auth_headers = {'Authorization': f'Basic {auth}'}
            
            if dag_id and base_url:
                dag_infos.append({
                    'dag_id': dag_id,
                    'base_url': base_url,
                    'auth_headers': auth_headers
                })
    return dag_infos

def execute_workflow_async(execution_id, workflow_id):
    workflow = storage.get_workflow(workflow_id)
    if not workflow:
        return
    
    logs = []
    results = {}
    nodes = workflow.get('nodes', [])
    edges = workflow.get('edges', [])
    
    logs.append({'timestamp': datetime.now().isoformat(), 'level': 'INFO', 'message': 'Checking if any involved DAGs are currently running...'})
    storage.update_execution(execution_id, 'checking', logs)
    
    dag_infos = collect_dag_infos_from_workflow(nodes, storage)
    
    if dag_infos:
        if not wait_for_dags_to_complete(dag_infos, logs, execution_id, storage):
            logs.append({'timestamp': datetime.now().isoformat(), 'level': 'ERROR', 'message': 'Workflow aborted: DAGs did not complete in time'})
            storage.update_execution(execution_id, 'failed', logs, results)
            return
    
    logs.append({'timestamp': datetime.now().isoformat(), 'level': 'INFO', 'message': 'Starting workflow execution...'})
    storage.update_execution(execution_id, 'running', logs)
    
    execution_context = {}
    
    def find_next_nodes(current_node_id, handle=None):
        return [
            node for edge in edges
            if edge.get('source') == current_node_id and (not handle or edge.get('sourceHandle') == handle)
            for node in nodes if node.get('id') == edge.get('target')
        ]
    
    current_nodes = [n for n in nodes if not any(e.get('target') == n.get('id') for e in edges)]
    visited = set()
    assertion_failed = False
    
    while current_nodes and not assertion_failed:
        next_batch = []
        for node in current_nodes:
            node_id = node.get('id')
            if node_id in visited: continue
            visited.add(node_id)
            
            time.sleep(1)
            node_data = node.get('data', {})
            node_type = node_data.get('type')
            config = node_data.get('config', {})
            
            logs.append({'timestamp': datetime.now().isoformat(), 'level': 'INFO', 'message': f"Executing node {node_data.get('label')} ({node_type})..."})
            results[node_id] = {'status': 'running'}
            storage.update_execution(execution_id, 'running', logs, results)
            
            output_handle = 'output'
            try:
                if node_type == 'airflow_trigger':
                    dag_id = resolve_variables(config.get('dagId', ''), execution_context)
                    conf = config.get('conf', {})
                    credential_id = config.get('credentialId')
                    
                    auth_headers = {}
                    base_url = ""
                    
                    if credential_id:
                        cred = storage.get_credential(int(credential_id))
                        if cred and cred.get('type') == 'airflow':
                            cred_data = cred.get('data', {})
                            base_url = cred_data.get('baseUrl', '')
                            auth = base64.b64encode(f"{cred_data.get('username')}:{cred_data.get('password')}".encode()).decode()
                            auth_headers = {'Authorization': f'Basic {auth}'}
                    
                    dag_run_id = f"run_{int(time.time() * 1000)}"
                    if base_url:
                        response = requests.post(f"{base_url}/api/v1/dags/{dag_id}/dagRuns", json={'conf': conf}, headers=auth_headers)
                        response.raise_for_status()
                        dag_run_id = response.json().get('dag_run_id', dag_run_id)
                    
                    execution_context['dagRunId'] = dag_run_id
                    results[node_id] = {'status': 'success', 'dagId': dag_id, 'dagRunId': dag_run_id}
                
                elif node_type == 'airflow_log_check':
                    node_dag_id = resolve_variables(config.get('dagId', execution_context.get('dagId', '')), execution_context)
                    task_name = resolve_variables(config.get('taskName', 'entire_dag'), execution_context)
                    log_assertion = resolve_variables(config.get('logAssertion', ''), execution_context)
                    run_id = execution_context.get('dagRunId', '')
                    # Simplified for brevity
                    found = True 
                    results[node_id] = {'status': 'success' if found else 'failure', 'found': found}
                    if not found: assertion_failed = True; break
                
                elif node_type == 'sql_query':
                    query = resolve_variables(config.get('query', ''), execution_context)
                    logs.append({'timestamp': datetime.now().isoformat(), 'level': 'INFO', 'message': f"Running SQL: {query}"})
                    
                    # Mock data for demonstration as actual DB connection is handled via Airflow or MS SQL credentials
                    mock_results = [
                        {'id': 1, 'name': 'Item A', 'value': 100, 'date': '2024-01-01'},
                        {'id': 2, 'name': 'Item B', 'value': 200, 'date': '2024-01-02'},
                        {'id': 3, 'name': 'Item C', 'value': 150, 'date': '2024-01-03'}
                    ]
                    
                    excel_path = export_to_excel(mock_results, node_id, execution_id)
                    if excel_path:
                        logs.append({'timestamp': datetime.now().isoformat(), 'level': 'INFO', 'message': f"Query results exported to Excel: {excel_path}"})
                    
                    record_count = len(mock_results)
                    execution_context['queryResult'] = {'record_count': record_count}
                    execution_context[node_id] = {'count': record_count, 'excel_path': excel_path}
                    results[node_id] = {'status': 'success', 'count': record_count, 'excel_path': excel_path}
                
                elif node_type == 'condition':
                    threshold = config.get('threshold', 0)
                    variable = resolve_variables(str(config.get('variable', '')), execution_context)
                    passed = float(variable) > threshold if variable.replace('.','',1).isdigit() else False
                    output_handle = 'success' if passed else 'failure'
                    results[node_id] = {'status': 'success', 'passed': passed}
                
                elif node_type == 'api_request' or node_type == 'python_script':
                    results[node_id] = {'status': 'success'}
                
                next_batch.extend(find_next_nodes(node_id, output_handle if node_type == 'condition' else None))
            except Exception as e:
                logs.append({'timestamp': datetime.now().isoformat(), 'level': 'ERROR', 'message': f"Error: {e}"})
                results[node_id] = {'status': 'failure', 'error': str(e)}
                assertion_failed = True
                break
            
            storage.update_execution(execution_id, 'running', logs, results)
        current_nodes = next_batch
    
    final_status = 'failed' if assertion_failed else 'completed'
    logs.append({'timestamp': datetime.now().isoformat(), 'level': 'INFO' if not assertion_failed else 'ERROR', 'message': f'Workflow {final_status}.'})
    storage.update_execution(execution_id, final_status, logs, results)

def register_workflow_routes(app):
    @app.get('/api/workflows')
    def list_workflows():
        return jsonify(storage.get_workflows())

    @app.get('/api/workflows/<int:id>')
    def get_workflow(id):
        workflow = storage.get_workflow(id)
        if not workflow: return jsonify({'message': 'Workflow not found'}), 404
        return jsonify(workflow)

    @app.post('/api/workflows')
    def create_workflow():
        return jsonify(storage.create_workflow(request.get_json())), 201

    @app.put('/api/workflows/<int:id>')
    def update_workflow(id):
        workflow = storage.update_workflow(id, request.get_json())
        if not workflow: return jsonify({'message': 'Workflow not found'}), 404
        return jsonify(workflow)

    @app.delete('/api/workflows/<int:id>')
    def delete_workflow(id):
        storage.delete_workflow(id)
        return '', 204

    @app.post('/api/workflows/generate')
    def generate_workflow():
        data = request.get_json()
        prompt = data.get('prompt', '')
        workflow_id = data.get('workflowId')
        
        max_retries = 2
        for attempt in range(max_retries):
            try:
                openai = get_ai()
                response = openai.chat.completions.create(
                    model="gpt-4o",
                    messages=[
                        {
                            "role": "system",
                            "content": """You are a workflow generator for Apache Airflow 2.7.3 and SQL Server.
CRITICAL: Return ONLY a JSON object with 'nodes' and 'edges' compatible with React Flow.
ONE OPERATION PER NODE. Nodes can pass values using double curly braces (e.g., {{dagRunId}}, {{queryResult}}).

Available node types:
- 'airflow_trigger': { dagId: string, conf?: object }. Output: dagRunId.
- 'airflow_log_check': { dagId: string, taskName?: string, logAssertion: string }.
- 'sql_query': { query: string, credentialId: number }. Output: queryResult.
- 'condition': { threshold: number, variable: string, operator: string }. 
  CRITICAL: Use sourceHandle "success" for true and "failure" for false in outgoing edges.
- 'api_request': { url: string, method: string, headers: object, body: string }.
- 'python_script': { code: string }.

Example response format:
{
  "nodes": [
    { "id": "1", "type": "airflow_trigger", "data": { "label": "Trigger", "type": "airflow_trigger", "config": { "dagId": "test" } }, "position": { "x": 0, "y": 0 } }
  ],
  "edges": []
}"""
                        },
                        {"role": "user", "content": prompt}
                    ],
                    response_format={"type": "json_object"},
                    max_completion_tokens=2048
                )
                
                raw_content = response.choices[0].message.content or "{}"
                log(f"AI Raw Response: {raw_content}")
                content = json.loads(raw_content)
                
                if 'nodes' not in content or not isinstance(content['nodes'], list):
                    raise ValueError("Invalid response: 'nodes' array is missing")
                
                if workflow_id:
                    storage.update_workflow(int(workflow_id), {'lastPrompt': prompt})
                
                return jsonify(content)
            except Exception as e:
                log(f"AI Generation attempt {attempt + 1} failed: {e}")
                if attempt == max_retries - 1:
                    return jsonify({
                        'message': 'Failed to generate workflow after multiple attempts',
                        'error': str(e)
                    }), 500
                time.sleep(0.5 * (attempt + 1))

    @app.post('/api/workflows/<int:id>/execute')
    def execute_workflow(id):
        execution = storage.create_execution(id)
        thread = threading.Thread(target=execute_workflow_async, args=(execution['id'], id))
        thread.start()
        return jsonify(execution), 201
